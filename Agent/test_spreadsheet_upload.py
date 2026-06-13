#!/usr/bin/env python3
"""Spreadsheet upload framework smoke tests.

Uses CSV so the test does not depend on optional Excel runtime packages.
"""

import tempfile
from pathlib import Path

import numpy as np
import faiss
import agents.knowledge_agent as knowledge_agent_module
import agents.reviewer_agent as reviewer_agent_module
import document_parser
from agents.orchestrator import AgentOrchestrator, ContextPacket
from agents.knowledge_agent import KnowledgeAgent
from agents.reviewer_agent import ReviewerAgent
from knowledge_base import KnowledgeBase
from knowledge_manifest import KnowledgeIngestionManifest
from spreadsheet_auditor import SpreadsheetFactAuditor
from spreadsheet_store import SpreadsheetStore, build_spreadsheet_validation, parse_spreadsheet
from upload_manager import UploadManager
from vector_map import build_vector_map


class FakeKnowledgeBase:
    def __init__(self, index_dir):
        self.index_dir = Path(index_dir)
        self.index_dir.mkdir(parents=True, exist_ok=True)
        self.metadatas = []
        self.documents = []

    def add_documents(self, documents):
        self.documents.extend(documents)
        self.metadatas.extend(doc.get("metadata", {}) for doc in documents)


class FakeEmbeddingModel:
    def encode(self, texts, **kwargs):
        vectors = []
        for text in texts:
            base = float((len(text) % 7) + 1)
            vectors.append([base, base / 10])
        return np.asarray(vectors, dtype=np.float32)


def test_csv_parse_and_store():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_store_") as tmp:
        csv_path = Path(tmp) / "产业报表.csv"
        csv_path.write_text(
            "年份,产业类别,产值,同比增速\n"
            "2024,人工智能,3600亿元,21.1%\n"
            "2025,人工智能,预计超4000亿元,待确认\n",
            encoding="utf-8",
        )

        rows = parse_spreadsheet(csv_path)
        assert len(rows) == 2
        assert rows[0].sheet_name == "CSV"
        assert rows[0].row_number == 2
        assert rows[0].values["产值"] == "3600亿元"

        store = SpreadsheetStore(Path(tmp) / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_csv",
            filename="产业报表.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=rows,
        )
        matches = store.find_cells(value="3600亿元", column_name="产值")
        assert len(matches) == 1
        assert matches[0]["sheet_name"] == "CSV"
        assert matches[0]["row_number"] == 2

        files = store.list_files()
        assert files[0]["validation"]["row_count"] == 2
        assert files[0]["validation"]["sheet_count"] == 1
        assert files[0]["validation"]["ok"] is True


def test_xls_extension_is_supported_with_clear_optional_dependency():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_xls_") as tmp:
        xls_path = Path(tmp) / "旧版报表.xls"
        xls_path.write_bytes(b"not-a-real-xls")

        manager = UploadManager(str(Path(tmp) / "uploads"))

        class FakeFile:
            filename = "旧版报表.xls"

            def __init__(self):
                self._pos = 0

            def seek(self, pos, whence=0):
                self._pos = pos

            def tell(self):
                return 10

        ok, message = manager.validate_file(FakeFile())
        assert ok is True

        try:
            parse_spreadsheet(xls_path)
        except Exception as exc:
            assert "xlrd" in str(exc) or "Unsupported format" in str(exc) or "not supported" in str(exc)


def test_csv_knowledge_upload_dual_track():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_upload_") as tmp:
        tmp_path = Path(tmp)
        upload_manager = UploadManager(str(tmp_path / "uploads"))
        kb = FakeKnowledgeBase(tmp_path / "knowledge_base")

        temp_file = tmp_path / "source.csv"
        temp_file.write_text(
            "年份,产业类别,产值,同比增速\n"
            "2024,人工智能,3600亿元,21.1%\n",
            encoding="utf-8",
        )

        result = upload_manager.process_knowledge_upload(
            file_path=str(temp_file),
            filename="产业报表.csv",
            category="公共资料",
            user_id="tester",
            knowledge_base=kb,
        )

        assert result["success"] is True
        assert result["chunks"] == 1
        assert Path(result["file_id"]).exists()
        assert kb.documents

        metadata = kb.documents[0]["metadata"]
        assert metadata["source_type"] == "spreadsheet"
        assert metadata["sheet_name"] == "CSV"
        assert metadata["row_start"] == 2
        assert metadata["column_headers"] == ["年份", "产业类别", "产值", "同比增速"]

        store = SpreadsheetStore(tmp_path / "knowledge_base" / "spreadsheets.sqlite")
        rows = store.get_rows_by_source(result["content_hash"])
        assert len(rows) == 1
        assert rows[0]["values"]["产值"] == "3600亿元"
        assert store.find_cells(value="3600亿元", content_hash=result["content_hash"])


def test_spreadsheet_query_respects_access_filter():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_query_") as tmp:
        public_csv = Path(tmp) / "公开报表.csv"
        public_csv.write_text("年份,产值\n2024,3600亿元\n", encoding="utf-8")
        restricted_csv = Path(tmp) / "部门报表.csv"
        restricted_csv.write_text("年份,产值\n2024,9000亿元\n", encoding="utf-8")

        store = SpreadsheetStore(Path(tmp) / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="public_hash",
            filename="公开报表.csv",
            source_path=str(public_csv),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=parse_spreadsheet(public_csv),
        )
        store.upsert_file_rows(
            content_hash="restricted_hash",
            filename="部门报表.csv",
            source_path=str(restricted_csv),
            category="财务部",
            access_level="restricted",
            department="财务部",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=parse_spreadsheet(restricted_csv),
        )

        public_only = store.query_rows(keyword="2024", access_filter={"access_level": ["public"]})
        assert len(public_only) == 1
        assert public_only[0]["filename"] == "公开报表.csv"

        finance = store.query_rows(
            keyword="2024",
            access_filter={"access_level": ["public", "restricted"], "department": ["财务部", ""]},
        )
        assert len(finance) == 2
        assert store.find_cells(
            value="9000亿元",
            access_filter={"access_level": ["public", "restricted"], "department": ["财务部", ""]},
        )
        assert not store.find_cells(value="9000亿元", access_filter={"access_level": ["public"]})


def test_knowledge_manifest_tracks_upload_consistency():
    with tempfile.TemporaryDirectory(prefix="knowledge_manifest_") as tmp:
        tmp_path = Path(tmp)
        upload_manager = UploadManager(str(tmp_path / "uploads"))
        kb = FakeKnowledgeBase(tmp_path / "knowledge_base")

        temp_file = tmp_path / "source.csv"
        temp_file.write_text(
            "年份,产业类别,产值,同比增速\n"
            "2024,人工智能,3600亿元,21.1%\n",
            encoding="utf-8",
        )

        result = upload_manager.process_knowledge_upload(
            file_path=str(temp_file),
            filename="产业报表.csv",
            category="公共资料",
            user_id="tester",
            knowledge_base=kb,
        )
        assert result["success"] is True

        manifest = KnowledgeIngestionManifest(tmp_path / "knowledge_base" / "ingestion_manifest.sqlite")
        record = manifest.get_record(result["content_hash"])
        assert record["status"] == "completed"
        assert record["chunk_count"] == 1
        assert record["spreadsheet_row_count"] == 1

        report = manifest.consistency_report(
            kb,
            tmp_path / "knowledge_base" / "spreadsheets.sqlite",
        )
        assert report["ok"] is True
        assert report["issue_count"] == 0

        Path(result["file_id"]).unlink()
        broken = manifest.consistency_report(
            kb,
            tmp_path / "knowledge_base" / "spreadsheets.sqlite",
        )
        assert broken["ok"] is False
        assert any("原文件不存在" in issue for issue in broken["issues"][0]["issues"])


def test_knowledge_admin_vector_replace_and_metadata_update():
    with tempfile.TemporaryDirectory(prefix="knowledge_admin_vectors_") as tmp:
        kb = KnowledgeBase(Path(tmp) / "knowledge_base", lazy_load=True)
        kb.dim = 2
        kb.model = FakeEmbeddingModel()
        kb.index = faiss.IndexFlatIP(2)
        kb.texts = ["旧文件A", "旧文件B"]
        kb.metadatas = [
            {"content_hash": "hash_a", "category": "公共资料", "access_level": "public", "department": ""},
            {"content_hash": "hash_b", "category": "公共资料", "access_level": "public", "department": ""},
        ]

        removed = kb.replace_by_content_hash("hash_a", [{
            "content": "新文件A片段",
            "metadata": {
                "content_hash": "hash_a",
                "category": "公共资料",
                "access_level": "public",
                "department": "",
            },
        }])
        assert removed == 1
        assert len(kb.texts) == 2
        assert sum(1 for meta in kb.metadatas if meta["content_hash"] == "hash_a") == 1
        assert kb.index.ntotal == 2

        updated = kb.update_metadata_by_content_hash("hash_a", {
            "category": "财务部",
            "access_level": "restricted",
            "department": "财务部",
        })
        assert updated == 1
        meta = next(meta for meta in kb.metadatas if meta["content_hash"] == "hash_a")
        assert meta["category"] == "财务部"
        assert meta["access_level"] == "restricted"
        assert meta["department"] == "财务部"

        deleted = kb.delete_by_content_hash("hash_a")
        assert deleted == 1
        assert all(meta["content_hash"] != "hash_a" for meta in kb.metadatas)
        assert kb.index.ntotal == 1


def test_manifest_archive_and_spreadsheet_metadata_sync():
    with tempfile.TemporaryDirectory(prefix="knowledge_admin_manifest_") as tmp:
        tmp_path = Path(tmp)
        csv_path = tmp_path / "source.csv"
        csv_path.write_text("年份,产值\n2024,3600亿元\n", encoding="utf-8")
        archived_path = tmp_path / "archive" / "source.csv"
        archived_path.parent.mkdir()
        archived_path.write_text(csv_path.read_text(encoding="utf-8"), encoding="utf-8")

        manifest = KnowledgeIngestionManifest(tmp_path / "ingestion_manifest.sqlite")
        manifest.record_started(
            content_hash="hash_admin",
            filename="source.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            source_type="spreadsheet",
            parser_type="spreadsheet_row",
            chunk_count=1,
        )
        manifest.mark_structured_indexed("hash_admin", 1)
        manifest.mark_vector_indexed("hash_admin")

        store = SpreadsheetStore(tmp_path / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_admin",
            filename="source.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=parse_spreadsheet(csv_path),
        )

        manifest.update_metadata(
            "hash_admin",
            category="财务部",
            access_level="restricted",
            department="财务部",
        )
        store.update_file_metadata(
            "hash_admin",
            category="财务部",
            access_level="restricted",
            department="财务部",
        )
        assert manifest.get_record("hash_admin")["department"] == "财务部"
        rows = store.query_rows(
            keyword="2024",
            access_filter={"access_level": ["public", "restricted"], "department": ["财务部", ""]},
        )
        assert rows[0]["category"] == "财务部"
        assert rows[0]["access_level"] == "restricted"

        store.delete_file("hash_admin")
        manifest.mark_archived("hash_admin", str(archived_path))
        report = manifest.consistency_report(
            type("KB", (), {"metadatas": []})(),
            tmp_path / "spreadsheets.sqlite",
        )
        assert report["ok"] is True
        assert report["recent"][0]["status"] == "archived"


def test_spreadsheet_validation_and_audit_log():
    with tempfile.TemporaryDirectory(prefix="knowledge_admin_audit_") as tmp:
        tmp_path = Path(tmp)
        csv_path = tmp_path / "quality.csv"
        csv_path.write_text(
            "年份,产值,备注\n"
            "2024,3600亿元,\n"
            "2024,3600亿元,\n",
            encoding="utf-8",
        )
        rows = parse_spreadsheet(csv_path)
        validation = build_spreadsheet_validation(rows)
        assert validation["row_count"] == 2
        assert validation["duplicate_rows"] == 1
        assert validation["warnings"]

        manifest = KnowledgeIngestionManifest(tmp_path / "manifest.sqlite")
        manifest.record_audit(
            content_hash="hash_audit",
            filename="quality.csv",
            action="reindex",
            actor_id="admin",
            actor_name="管理员",
            status="success",
            message="重建完成",
            backup_path=str(tmp_path / "backup"),
            before={"status": "completed"},
            after={"status": "completed", "chunk_count": 2},
        )
        audit = manifest.recent_audit()
        assert audit[0]["action"] == "reindex"
        assert audit[0]["actor_id"] == "admin"
        assert audit[0]["before"]["status"] == "completed"
        assert audit[0]["after"]["chunk_count"] == 2


def test_admin_operation_snapshots_restore_all_stores():
    with tempfile.TemporaryDirectory(prefix="knowledge_admin_rollback_") as tmp:
        tmp_path = Path(tmp)
        csv_path = tmp_path / "source.csv"
        csv_path.write_text("年份,产值\n2024,3600亿元\n", encoding="utf-8")

        kb = KnowledgeBase(tmp_path / "knowledge_base", lazy_load=True)
        kb.dim = 2
        kb.model = FakeEmbeddingModel()
        kb.index = faiss.IndexFlatIP(2)
        kb.add_documents([{
            "content": "旧片段",
            "metadata": {
                "content_hash": "hash_rollback",
                "category": "公共资料",
                "access_level": "public",
                "department": "",
            },
        }])

        manifest = KnowledgeIngestionManifest(tmp_path / "manifest.sqlite")
        manifest.record_started(
            content_hash="hash_rollback",
            filename="source.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            source_type="spreadsheet",
            parser_type="spreadsheet_row",
            chunk_count=1,
        )
        manifest.mark_structured_indexed("hash_rollback", 1)
        manifest.mark_vector_indexed("hash_rollback")

        store = SpreadsheetStore(tmp_path / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_rollback",
            filename="source.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=parse_spreadsheet(csv_path),
        )

        kb_snapshot = kb.snapshot_state()
        sheet_snapshot = store.snapshot_file("hash_rollback")
        manifest_snapshot = manifest.get_record("hash_rollback")

        kb.replace_by_content_hash("hash_rollback", [{
            "content": "错误新片段",
            "metadata": {
                "content_hash": "hash_rollback",
                "category": "财务部",
                "access_level": "restricted",
                "department": "财务部",
            },
        }])
        store.delete_file("hash_rollback")
        manifest.mark_archived("hash_rollback", str(tmp_path / "archive.csv"))

        kb.restore_state(kb_snapshot)
        store.restore_file_snapshot(sheet_snapshot)
        manifest.restore_record(manifest_snapshot)

        assert kb.texts == ["旧片段"]
        assert kb.metadatas[0]["access_level"] == "public"
        assert store.query_rows(keyword="3600亿元")[0]["access_level"] == "public"
        assert manifest.get_record("hash_rollback")["status"] == "completed"


def test_xlsx_realistic_report_shapes_when_available():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("SKIP_XLSX_REALISTIC_REPORT: openpyxl not available")
        return

    with tempfile.TemporaryDirectory(prefix="spreadsheet_xlsx_") as tmp:
        xlsx_path = Path(tmp) / "综合报表.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "产业数据"
        ws["A1"] = "智能知识库平台"
        ws.merge_cells("A1:D1")
        ws.append([])
        ws.append(["年份", "产业类别", "产值", "同比增速"])
        ws.append([2024, "人工智能", 3600, "21.1%"])
        ws.append([2025, "人工智能", "预计超4000亿元", "待确认"])

        ws2 = wb.create_sheet("项目数据")
        ws2.append(["年份", "项目名称", "数量"])
        ws2.append([2024, "重点项目", 18])
        wb.save(xlsx_path)

        rows = parse_spreadsheet(xlsx_path)
        assert len(rows) == 3
        industry = [row for row in rows if row.sheet_name == "产业数据"]
        projects = [row for row in rows if row.sheet_name == "项目数据"]
        assert len(industry) == 2
        assert len(projects) == 1
        assert industry[0].row_number == 4
        assert industry[0].values["产业类别"] == "人工智能"
        assert industry[0].values["产值"] == 3600
        assert projects[0].values["数量"] == 18


def test_xlsx_parser_ignores_styled_far_empty_columns_when_available():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        print("SKIP_XLSX_FAR_COLUMNS: openpyxl not available")
        return

    with tempfile.TemporaryDirectory(prefix="spreadsheet_xlsx_far_cols_") as tmp:
        xlsx_path = Path(tmp) / "远端空列.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["年份", "产值"])
        ws.append([2024, "3600亿元"])
        ws["XFD1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        wb.save(xlsx_path)

        rows = parse_spreadsheet(xlsx_path)
        assert len(rows) == 1
        assert rows[0].headers == ["年份", "产值"]
        assert rows[0].values == {"年份": 2024, "产值": "3600亿元"}


def test_spreadsheet_display_filename_is_used_in_row_text():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_display_name_") as tmp:
        csv_path = Path(tmp) / "tmp_upload.csv"
        csv_path.write_text("年份,产值\n2024,3600亿元\n", encoding="utf-8")

        rows = parse_spreadsheet(csv_path, display_filename="单位场地收支开票统计表(1).xlsx")
        assert rows[0].text.startswith("文件：单位场地收支开票统计表(1).xlsx")


def test_upload_manager_preserves_chinese_filename():
    manager = UploadManager()
    assert manager._safe_filename("2026年5月企业培训授课明细(1).xlsx", ".xlsx") == "2026年5月企业培训授课明细(1).xlsx"


def test_temp_upload_metadata_survives_new_manager_instance():
    with tempfile.TemporaryDirectory(prefix="temp_upload_shared_") as tmp:
        base_path = Path(tmp) / "uploads"
        source_path = base_path / "temp" / "worker_a_upload.txt"
        source_path.parent.mkdir(parents=True, exist_ok=True)
        source_path.write_text("跨 worker 临时上传内容", encoding="utf-8")

        worker_a = UploadManager(base_path=str(base_path))
        result = worker_a.process_temp_upload(
            file_path=str(source_path),
            filename="真实环境临时文件.txt",
            user_id="user_a",
        )

        worker_b = UploadManager(base_path=str(base_path))
        assert result["success"] is True
        assert worker_b.get_temp_content(result["file_id"], "user_a") == "跨 worker 临时上传内容"
        assert worker_b.get_temp_file_info(result["file_id"], "user_a")["filename"] == "真实环境临时文件.txt"
        assert worker_b.get_temp_content(result["file_id"], "user_b") is None


def test_pdf_parser_returns_page_level_items():
    class FakePage:
        def __init__(self, text):
            self._text = text

        def extract_text(self):
            return self._text

    class FakeReader:
        def __init__(self, _path):
            self.pages = [FakePage("第一页 经费管理"), FakePage(""), FakePage("第三页 审批程序")]

    import PyPDF2
    real_reader = PyPDF2.PdfReader
    PyPDF2.PdfReader = FakeReader
    try:
        pages = document_parser.parse_pdf_with_format(Path("fake.pdf"))
    finally:
        PyPDF2.PdfReader = real_reader

    assert len(pages) == 2
    assert pages[0]["page_start"] == 1
    assert pages[1]["page_start"] == 3
    assert pages[0]["parse_warnings"]


def test_text_upload_adds_heading_and_chunk_metadata():
    with tempfile.TemporaryDirectory(prefix="document_metadata_") as tmp:
        tmp_path = Path(tmp)
        md_path = tmp_path / "经费制度.md"
        md_path.write_text(
            "# 经费管理\n"
            "经费支出审批程序应当先提交申请，再完成部门审核。\n\n"
            "## 报销材料\n"
            "报销材料包括合同、发票和审批单。\n",
            encoding="utf-8",
        )

        manager = UploadManager(str(tmp_path / "uploads"))
        prepared = manager.build_knowledge_documents(
            file_path=str(md_path),
            filename="经费制度.md",
            category="公共资料",
            user_id="tester",
        )

        assert prepared["success"] is True
        assert prepared["parser_type"] == "text_structural"
        first_meta = prepared["documents"][0]["metadata"]
        assert first_meta["section_title"] == "经费管理"
        assert first_meta["heading_path"] == ["经费管理"]
        assert first_meta["chunk_text_hash"]
        assert "page_start" in first_meta


def test_bm25_uses_enhanced_document_fields():
    texts = ["正文没有目标词", "普通正文 经费", "第三份普通材料"]
    metadatas = [
        {
            "source": "a.md",
            "filename": "经费制度.md",
            "category": "公共资料",
            "doc_type": "参考材料",
            "source_type": "document",
            "section_title": "经费管理",
            "heading_path": ["经费管理"],
            "access_level": "public",
        },
        {
            "source": "b.md",
            "filename": "普通材料.md",
            "category": "公共资料",
            "doc_type": "参考材料",
            "source_type": "document",
            "section_title": "",
            "heading_path": [],
            "access_level": "public",
        },
        {
            "source": "c.md",
            "filename": "其他材料.md",
            "category": "公共资料",
            "doc_type": "参考材料",
            "source_type": "document",
            "section_title": "",
            "heading_path": [],
            "access_level": "public",
        },
    ]
    agent = KnowledgeAgent.__new__(KnowledgeAgent)
    agent.index_data = {"texts": texts, "metadatas": metadatas}
    tokenized = [
        list(knowledge_agent_module.jieba.cut(agent._bm25_document_text(text, metadatas[i])))
        for i, text in enumerate(texts)
    ]
    agent.bm25 = knowledge_agent_module.BM25Okapi(tokenized)

    results = agent._search_bm25("经费管理", access_filter=None, top_k=2)
    assert results[0]["filename"] == "经费制度.md"
    assert results[0]["section_title"] == "经费管理"


def test_spreadsheet_parser_keeps_summary_and_skips_invoice_footer():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_invoice_footer_") as tmp:
        csv_path = Path(tmp) / "授课明细.csv"
        csv_path.write_text(
            "使用日期,教室名称,课程/班级,费用项目,数量,单位,单价,总费用\n"
            "2026.05.21,501,企业培训,场地使用费,1,天,4000,2800\n"
            "总计（人民币：贰仟捌佰元整）,总计（人民币：贰仟捌佰元整）,总计（人民币：贰仟捌佰元整）,总计（人民币：贰仟捌佰元整）,,,,2800\n"
            "请核对确认，如核对无误，请将上述款项汇入以下账户，谢谢！,,,,,,,\n"
            "户名：智能知识库平台,账号：000000000000000,开户行：示例银行示例支行,,,,,\n",
            encoding="utf-8",
        )

        rows = parse_spreadsheet(csv_path)
        assert len(rows) == 2
        assert [row.row_type for row in rows] == ["data", "summary"]
        assert "请核对确认" not in "\n".join(row.text for row in rows)


def test_knowledge_context_uses_structured_spreadsheet_row():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_context_") as tmp:
        tmp_path = Path(tmp)
        csv_path = tmp_path / "产业报表.csv"
        csv_path.write_text(
            "年份,产业类别,产值,同比增速\n"
            "2024,人工智能,3600亿元,21.1%\n",
            encoding="utf-8",
        )
        rows = parse_spreadsheet(csv_path)
        store = SpreadsheetStore(tmp_path / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_csv",
            filename="产业报表.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=rows,
        )

        original_db_path = knowledge_agent_module.SPREADSHEET_DB_PATH
        knowledge_agent_module.SPREADSHEET_DB_PATH = str(tmp_path / "spreadsheets.sqlite")
        try:
            agent = KnowledgeAgent.__new__(KnowledgeAgent)
            results = [{
                "text": "向量片段可能被压缩",
                "source": str(csv_path),
                "filename": "产业报表.csv",
                "source_type": "spreadsheet",
                "content_hash": "hash_csv",
                "sheet_name": "CSV",
                "row_start": 2,
                "row_end": 2,
                "chunk_index": 0,
                "total_chunks": 1,
                "similarity": 0.9,
            }]

            hydrated = agent._hydrate_spreadsheet_rows(results)
            context = agent._build_context(hydrated)
            assert "3600亿元" in context
            assert "Sheet: CSV，行: 2" in context
            assert "不可自行换算、补全或推测" in context
            assert "向量片段可能被压缩" not in context
        finally:
            knowledge_agent_module.SPREADSHEET_DB_PATH = original_db_path


def test_knowledge_context_expands_spreadsheet_table_for_list_request():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_table_context_") as tmp:
        tmp_path = Path(tmp)
        csv_path = tmp_path / "附件：场地使用收费表.csv"
        csv_path.write_text(
            "序号,楼层,门牌号,名称,可容纳人数,计费方式,金额,备注\n"
            "序号,楼层,门牌号,名称,可容纳人数,计费方式,金额,备注\n"
            "1,22F,2215,智慧教室,100,1天,6000,智慧屏\n"
            "2,22F,2216,智慧教室,50,1天,5000,音响系统\n"
            "3,22F,2206,多功能报告厅,150,1天,40000,大屏\n"
            "4,宝安院区,宝安院区,宝安院区,宝安院区,宝安院区,宝安院区,宝安院区\n"
            "5,前海院区,22F,共享工作室,11,300,30,水电网络\n",
            encoding="utf-8",
        )
        rows = parse_spreadsheet(csv_path)
        store = SpreadsheetStore(tmp_path / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_fee",
            filename="附件：场地使用收费表.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-06-11T00:00:00",
            rows=rows,
        )

        original_db_path = knowledge_agent_module.SPREADSHEET_DB_PATH
        knowledge_agent_module.SPREADSHEET_DB_PATH = str(tmp_path / "spreadsheets.sqlite")
        try:
            agent = KnowledgeAgent.__new__(KnowledgeAgent)
            results = [{
                "text": "只命中了表头",
                "source": str(csv_path),
                "filename": "附件：场地使用收费表.csv",
                "source_type": "spreadsheet",
                "content_hash": "hash_fee",
                "sheet_name": "CSV",
                "row_start": 2,
                "row_end": 2,
                "row_type": "data",
                "chunk_index": 0,
                "total_chunks": 4,
                "similarity": 0.7,
            }]

            hydrated = agent._hydrate_spreadsheet_rows(results, user_request="场地使用收费表有哪些教室和收费")
            context = agent._build_context(hydrated)

            assert "2215" in context
            assert "智慧教室" in context
            assert "6000" in context
            assert "多功能报告厅" in context
            assert "40000" in context
            assert "宝安院区" not in context
            assert "共享工作室" not in context
            assert "只命中了表头" not in context
        finally:
            knowledge_agent_module.SPREADSHEET_DB_PATH = original_db_path


def test_knowledge_context_prefers_latest_equivalent_spreadsheet_table():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_latest_fee_") as tmp:
        tmp_path = Path(tmp)
        old_csv = tmp_path / "20260604_附件：场地使用收费表.csv"
        old_csv.write_text(
            "序号,楼层,门牌号,名称,可容纳人数,计费方式,金额,备注\n"
            "1,22F,2215,智慧教室,100,1天,6000,旧版价格\n",
            encoding="utf-8",
        )
        new_csv = tmp_path / "20260611_附件：场地使用收费表（0420版本)(2)(1).csv"
        new_csv.write_text(
            "序号,楼层,门牌号,名称,可容纳人数,计费方式,金额,备注\n"
            "1,22F,2215,智慧教室,100,1天,6600,新版价格\n",
            encoding="utf-8",
        )

        store = SpreadsheetStore(tmp_path / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_fee_old",
            filename="附件：场地使用收费表.csv",
            source_path=str(old_csv),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-06-04T00:00:00",
            rows=parse_spreadsheet(old_csv),
        )
        store.upsert_file_rows(
            content_hash="hash_fee_new",
            filename="附件：场地使用收费表（0420版本)(2)(1).csv",
            source_path=str(new_csv),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-06-11T00:00:00",
            rows=parse_spreadsheet(new_csv),
        )

        original_db_path = knowledge_agent_module.SPREADSHEET_DB_PATH
        knowledge_agent_module.SPREADSHEET_DB_PATH = str(tmp_path / "spreadsheets.sqlite")
        try:
            agent = KnowledgeAgent.__new__(KnowledgeAgent)
            results = [
                {
                    "text": "旧版向量片段",
                    "source": str(old_csv),
                    "source_path": str(old_csv),
                    "filename": "附件：场地使用收费表.csv",
                    "source_type": "spreadsheet",
                    "content_hash": "hash_fee_old",
                    "sheet_name": "CSV",
                    "row_start": 2,
                    "row_end": 2,
                    "row_type": "data",
                    "chunk_index": 0,
                    "total_chunks": 1,
                    "similarity": 0.98,
                    "uploaded_at": "2026-06-04T00:00:00",
                },
                {
                    "text": "新版向量片段",
                    "source": str(new_csv),
                    "source_path": str(new_csv),
                    "filename": "附件：场地使用收费表（0420版本)(2)(1).csv",
                    "source_type": "spreadsheet",
                    "content_hash": "hash_fee_new",
                    "sheet_name": "CSV",
                    "row_start": 2,
                    "row_end": 2,
                    "row_type": "data",
                    "chunk_index": 0,
                    "total_chunks": 1,
                    "similarity": 0.92,
                    "uploaded_at": "2026-06-11T00:00:00",
                },
            ]

            hydrated = agent._hydrate_spreadsheet_rows(
                results,
                user_request="场地使用收费表 教室 会议室 报告厅 贵宾厅 门牌号 可容纳人数 计费方式 金额 收费标准",
            )
            context = agent._build_context(hydrated)

            assert "附件：场地使用收费表（0420版本)(2)(1).csv" in context
            assert "6600" in context
            assert "新版价格" in context
            assert "旧版价格" not in context
            assert "6000" not in context
        finally:
            knowledge_agent_module.SPREADSHEET_DB_PATH = original_db_path


def test_orchestrator_source_details_include_sheet_row():
    ctx = ContextPacket(user_request="查询产业报表")
    ctx.knowledge_sources = [{
        "filename": "产业报表.csv",
        "source_type": "spreadsheet",
        "source_path": "/tmp/产业报表.csv",
        "category": "公共资料",
        "sheet_name": "CSV",
        "row_start": 2,
        "row_end": 2,
        "column_headers": ["年份", "产业类别", "产值", "同比增速"],
        "chunk_index": 0,
        "total_chunks": 1,
    }]

    orchestrator = AgentOrchestrator.__new__(AgentOrchestrator)
    details = orchestrator._source_details(ctx)
    assert details[0]["source_type"] == "spreadsheet"
    assert details[0]["sheet_name"] == "CSV"
    assert details[0]["row_start"] == 2
    assert "产值" in details[0]["column_headers"]


def test_vector_map_projection_keeps_file_metadata():
    embeddings = np.array([
        [1.0, 0.0, 0.0, 0.2],
        [0.9, 0.1, 0.0, 0.1],
        [0.0, 1.0, 0.2, 0.0],
        [0.0, 0.9, 0.1, 0.1],
    ], dtype=np.float32)
    metadatas = [
        {"filename": "制度.docx", "category": "制度", "source_type": "document", "chunk_index": 0, "total_chunks": 2},
        {"filename": "制度.docx", "category": "制度", "source_type": "document", "chunk_index": 1, "total_chunks": 2},
        {"filename": "报表.xlsx", "category": "公共资料", "source_type": "spreadsheet", "sheet_name": "Sheet1", "row_start": 2},
        {"filename": "报表.xlsx", "category": "公共资料", "source_type": "spreadsheet", "sheet_name": "Sheet1", "row_start": 3},
    ]
    texts = ["制度片段一", "制度片段二", "产值：3600亿元", "产值：4000亿元"]
    payload = build_vector_map(embeddings, metadatas, texts, [0, 1, 2, 3])

    assert payload["ok"] is True
    assert payload["point_count"] == 4
    assert payload["file_count"] == 2
    assert len(payload["explained_variance"]) == 3
    assert all(-1.0001 <= point["x"] <= 1.0001 for point in payload["points"])
    spreadsheet_point = next(point for point in payload["points"] if point["source_type"] == "spreadsheet")
    assert spreadsheet_point["sheet_name"] == "Sheet1"
    assert spreadsheet_point["row_start"] == 2


def test_spreadsheet_fact_auditor_blocks_unverified_number():
    auditor = SpreadsheetFactAuditor()
    evidence_items = [{
        "source_type": "spreadsheet",
        "content_hash": "hash_csv",
        "spreadsheet_values": [{
            "年份": "2024",
            "产业类别": "人工智能",
            "产值": "3600亿元",
            "同比增速": "21.1%",
        }],
    }]

    ok = auditor.audit("2024年人工智能产值为3600亿元，同比增长21.1%。", evidence_items)
    assert ok["passed"] is True
    assert "3600亿元" in ok["verified_claims"]

    bad = auditor.audit("2024年人工智能产值为5000亿元，同比增长21.1%。", evidence_items)
    assert bad["passed"] is False
    assert "5000亿元" in bad["unverified_claims"]


def test_reviewer_spreadsheet_fact_check_requires_exact_evidence():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_review_") as tmp:
        tmp_path = Path(tmp)
        csv_path = tmp_path / "产业报表.csv"
        csv_path.write_text(
            "年份,产业类别,产值,同比增速\n"
            "2024,人工智能,3600亿元,21.1%\n",
            encoding="utf-8",
        )
        rows = parse_spreadsheet(csv_path)
        store = SpreadsheetStore(tmp_path / "spreadsheets.sqlite")
        store.upsert_file_rows(
            content_hash="hash_csv",
            filename="产业报表.csv",
            source_path=str(csv_path),
            category="公共资料",
            access_level="public",
            department="",
            uploaded_by="tester",
            uploaded_at="2026-05-28T00:00:00",
            rows=rows,
        )

        original_db_path = reviewer_agent_module.SPREADSHEET_DB_PATH
        reviewer_agent_module.SPREADSHEET_DB_PATH = str(tmp_path / "spreadsheets.sqlite")
        try:
            reviewer = ReviewerAgent.__new__(ReviewerAgent)
            reviewer.name = "Reviewer"
            content = (
                "关于人工智能产业发展的报告\n\n"
                "有关单位：\n"
                "一、基本情况\n"
                "根据产业报表，2024年人工智能产值为5000亿元，同比增长21.1%。"
                "该数据将作为后续工作研判的重要依据。\n"
                "二、下一步工作\n"
                "持续加强数据审核和材料复核，确保引用口径一致、来源清晰。\n\n"
                "示例高校\n"
                "2026年5月28日"
            )
            result = reviewer.process({
                "document_content": content,
                "document_type": "报告",
                "task_type": "公文生成",
                "source_filenames": ["产业报表.csv"],
                "evidence_items": [{
                    "type": "spreadsheet",
                    "title": "产业报表.csv",
                    "source_type": "spreadsheet",
                    "content_hash": "hash_csv",
                    "sheet_name": "CSV",
                    "row_start": 2,
                    "row_end": 2,
                }],
            })

            meta = result.metadata
            assert meta["needs_revision"] is True
            assert meta["fact_check"]["passed"] is False
            assert any("5000亿元" in issue for issue in meta["fact_check"]["issues"])
        finally:
            reviewer_agent_module.SPREADSHEET_DB_PATH = original_db_path


if __name__ == "__main__":
    test_csv_parse_and_store()
    test_xls_extension_is_supported_with_clear_optional_dependency()
    test_csv_knowledge_upload_dual_track()
    test_spreadsheet_query_respects_access_filter()
    test_knowledge_manifest_tracks_upload_consistency()
    test_knowledge_admin_vector_replace_and_metadata_update()
    test_manifest_archive_and_spreadsheet_metadata_sync()
    test_spreadsheet_validation_and_audit_log()
    test_admin_operation_snapshots_restore_all_stores()
    test_xlsx_realistic_report_shapes_when_available()
    test_xlsx_parser_ignores_styled_far_empty_columns_when_available()
    test_spreadsheet_display_filename_is_used_in_row_text()
    test_upload_manager_preserves_chinese_filename()
    test_pdf_parser_returns_page_level_items()
    test_text_upload_adds_heading_and_chunk_metadata()
    test_bm25_uses_enhanced_document_fields()
    test_spreadsheet_parser_keeps_summary_and_skips_invoice_footer()
    test_knowledge_context_uses_structured_spreadsheet_row()
    test_orchestrator_source_details_include_sheet_row()
    test_vector_map_projection_keeps_file_metadata()
    test_spreadsheet_fact_auditor_blocks_unverified_number()
    test_reviewer_spreadsheet_fact_check_requires_exact_evidence()
    print("SPREADSHEET_UPLOAD_PASS")
