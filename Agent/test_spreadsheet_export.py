#!/usr/bin/env python3
"""Excel export helper smoke tests."""

from io import BytesIO

from openpyxl import load_workbook

from spreadsheet_export import (
    extract_markdown_tables,
    workbook_from_structured_rows,
    workbook_from_text,
    workbook_to_bytes,
)


def _load_workbook(workbook):
    return load_workbook(BytesIO(workbook_to_bytes(workbook)), data_only=False)


def test_markdown_table_exports_to_sheet():
    content = """
| 年份 | 产值 | 备注 |
| --- | ---: | --- |
| 2025 | 4000亿元 | 待确认 |
| 2026 | =危险公式 | 预测 |
"""
    tables = extract_markdown_tables(content)
    assert len(tables) == 1
    assert tables[0][0] == ["年份", "产值", "备注"]

    loaded = _load_workbook(workbook_from_text(content))
    ws = loaded["表格1"]
    assert ws["A1"].value == "年份"
    assert ws["B2"].value == "4000亿元"
    assert ws["B3"].value == "'=危险公式"


def test_structured_rows_export_keeps_sheet_and_row_metadata():
    rows = [
        {
            "filename": "测试报表.xlsx",
            "sheet_name": "Sheet/一",
            "row_number": 2,
            "row_type": "data",
            "headers": ["部门", "金额"],
            "values": {"部门": "综合部", "金额": "1200"},
            "category": "公共资料",
            "access_level": "public",
            "department": "",
        },
        {
            "filename": "测试报表.xlsx",
            "sheet_name": "Sheet/一",
            "row_number": 3,
            "row_type": "summary",
            "headers": ["部门", "金额"],
            "values": {"部门": "合计", "金额": "+1200"},
            "category": "公共资料",
            "access_level": "public",
            "department": "",
        },
    ]

    loaded = _load_workbook(workbook_from_structured_rows(rows))
    assert "导出说明" in loaded.sheetnames
    assert "Sheet_一" in loaded.sheetnames
    ws = loaded["Sheet_一"]
    assert [ws["A1"].value, ws["B1"].value, ws["C1"].value, ws["D1"].value] == [
        "原始行号",
        "行类型",
        "部门",
        "金额",
    ]
    assert ws["A2"].value == 2
    assert ws["C2"].value == "综合部"
    assert ws["D3"].value == "'+1200"


if __name__ == "__main__":
    test_markdown_table_exports_to_sheet()
    test_structured_rows_export_keeps_sheet_and_row_metadata()
    print("PASS spreadsheet export tests")
