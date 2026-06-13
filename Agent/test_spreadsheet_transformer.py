#!/usr/bin/env python3
"""Spreadsheet natural-language transform smoke tests."""

from io import BytesIO
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from spreadsheet_transformer import (
    apply_operation,
    collect_headers,
    infer_operation_spec,
    transform_spreadsheet_file,
)
from spreadsheet_store import parse_spreadsheet


def test_infer_filter_sort_and_export_csv():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_transform_") as tmp:
        csv_path = Path(tmp) / "项目清单.csv"
        csv_path.write_text(
            "项目,状态,预算,截止日期\n"
            "A,进行中,8万,2026-06-10\n"
            "B,已完成,20万,2026-06-01\n"
            "C,进行中,15万,2026-06-05\n",
            encoding="utf-8",
        )

        rows = parse_spreadsheet(csv_path)
        headers = collect_headers(rows)
        spec = infer_operation_spec("筛选状态为进行中，按预算从高到低排序，导出前2条", headers)
        assert spec.filters[0].column == "状态"
        assert spec.filters[0].operator == "eq"
        assert spec.sorts[0].column == "预算"
        assert spec.sorts[0].direction == "desc"
        assert spec.limit == 2

        transformed = apply_operation(rows, spec)
        assert [row.values["项目"] for row in transformed] == ["C", "A"]

        result = transform_spreadsheet_file(
            str(csv_path),
            "项目清单.csv",
            "筛选状态为进行中，按预算从高到低排序，导出前2条",
        )
        assert result["success"] is True
        assert result["summary"]["original_count"] == 3
        assert result["summary"]["output_count"] == 2

        wb = load_workbook(BytesIO(result["content"]), data_only=False)
        ws = wb["处理结果"]
        assert ws["A1"].value == "来源Sheet"
        assert ws["C2"].value == "C"
        assert ws["E2"].value == "15万"
        assert "处理说明" in wb.sheetnames


def test_transform_rejects_unclear_instruction():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_transform_unclear_") as tmp:
        csv_path = Path(tmp) / "项目清单.csv"
        csv_path.write_text(
            "项目,状态,预算\n"
            "A,进行中,8万\n",
            encoding="utf-8",
        )

        result = transform_spreadsheet_file(str(csv_path), "项目清单.csv", "帮我处理一下")
        assert result["success"] is False
        assert "没有识别到明确" in result["message"]


def test_date_sort_uses_full_date_not_year_only():
    with tempfile.TemporaryDirectory(prefix="spreadsheet_transform_date_") as tmp:
        csv_path = Path(tmp) / "任务清单.csv"
        csv_path.write_text(
            "任务,截止日期\n"
            "A,2026-06-10\n"
            "B,2026-06-01\n"
            "C,2026-06-05\n",
            encoding="utf-8",
        )

        rows = parse_spreadsheet(csv_path)
        spec = infer_operation_spec("按截止日期由早到晚排序", collect_headers(rows))
        transformed = apply_operation(rows, spec)
        assert [row.values["任务"] for row in transformed] == ["B", "C", "A"]


if __name__ == "__main__":
    test_infer_filter_sort_and_export_csv()
    test_transform_rejects_unclear_instruction()
    test_date_sort_uses_full_date_not_year_only()
    print("PASS spreadsheet transformer tests")
