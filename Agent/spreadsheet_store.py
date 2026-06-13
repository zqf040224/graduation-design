"""
Spreadsheet ingestion and structured storage.

Excel/CSV files are treated as verifiable data assets:
1. row-level text chunks for vector retrieval
2. SQLite rows for exact lookup and answer validation
"""

import csv
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Union


SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}


@dataclass
class SpreadsheetRow:
    sheet_name: str
    row_number: int
    headers: List[str]
    values: Dict[str, Any]
    text: str
    row_type: str = "data"


def is_spreadsheet_file(file_path: Union[str, Path]) -> bool:
    return Path(file_path).suffix.lower() in SPREADSHEET_EXTENSIONS


def parse_spreadsheet(file_path: Union[str, Path], display_filename: str = None) -> List[SpreadsheetRow]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(path, display_filename=display_filename)
    if suffix == ".xls":
        return _parse_xls(path, display_filename=display_filename)
    if suffix == ".csv":
        return _parse_csv(path, display_filename=display_filename)
    raise ValueError(f"暂不支持的表格格式: {suffix}")


def _normalize_header(value: Any, index: int) -> str:
    text = _format_cell(value).strip()
    return text or f"未命名列{index + 1}"


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _row_to_text(filename: str, sheet_name: str, row_number: int,
                 values: Dict[str, Any], row_type: str = "data") -> str:
    parts = [
        f"文件：{filename}",
        f"Sheet：{sheet_name}",
        f"行号：{row_number}",
    ]
    if row_type != "data":
        parts.append(f"行类型：{row_type}")
    for header, value in values.items():
        cell = _format_cell(value)
        if cell:
            parts.append(f"{header}：{cell}")
    return "\n".join(parts)


FOOTER_PATTERNS = [
    "请核对确认",
    "请将上述款项汇入",
    "户名",
    "账号",
    "开户行",
    "谢谢",
    "智能知识库平台",
    "示例单位",
    "审核：",
    "制表人：",
]


def _row_text_blob(values: Dict[str, Any]) -> str:
    return "\n".join(_format_cell(value) for value in values.values() if _format_cell(value))


def _classify_row(values: Dict[str, Any]) -> str:
    """Classify spreadsheet rows so invoice footers do not pollute detail retrieval."""
    blob = _row_text_blob(values)
    compact = "".join(blob.split())
    if not compact:
        return "empty"

    if "总计" in compact or "合计" in compact:
        return "summary"

    if any(pattern in compact for pattern in FOOTER_PATTERNS):
        return "footer"

    if re.fullmatch(r"(?:19|20)\d{2}年\d{1,2}月\d{1,2}日", compact):
        return "footer"

    unique_values = {_format_cell(value) for value in values.values() if _format_cell(value)}
    if len(unique_values) == 1:
        only = next(iter(unique_values))
        if len(only) > 20 and (
            any(pattern in only for pattern in FOOTER_PATTERNS)
            or re.search(r"(?:19|20)\d{2}年\d{1,2}月\d{1,2}日", only)
        ):
            return "footer"

    return "data"


def _first_non_empty_row(rows: List[List[Any]]) -> Optional[int]:
    for idx, row in enumerate(rows):
        if any(_format_cell(value) for value in row):
            return idx
    return None


def _find_header_row(rows: List[List[Any]]) -> Optional[int]:
    """Find a likely header row, skipping report titles and blank preambles."""
    fallback = _first_non_empty_row(rows)
    for idx, row in enumerate(rows):
        cells = [_format_cell(value) for value in row]
        non_empty = [cell for cell in cells if cell]
        if len(non_empty) < 2:
            continue
        if len(set(non_empty)) < 2:
            continue
        following = rows[idx + 1:idx + 4]
        has_following_data = any(
            sum(1 for value in next_row if _format_cell(value)) >= 2
            for next_row in following
        )
        if has_following_data:
            return idx
    return fallback


def _parse_xlsx(path: Path, display_filename: str = None) -> List[SpreadsheetRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法解析 .xlsx 文件，请先安装依赖") from exc

    workbook = load_workbook(path, data_only=True, read_only=False)
    parsed: List[SpreadsheetRow] = []

    for sheet in workbook.worksheets:
        rows = _sheet_rows_with_merged_values(sheet)
        header_idx = _find_header_row(rows)
        if header_idx is None:
            continue

        headers = [_normalize_header(value, i) for i, value in enumerate(rows[header_idx])]
        for offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            values = {}
            for i, header in enumerate(headers):
                cell_value = row[i] if i < len(row) else None
                if _format_cell(cell_value):
                    values[header] = cell_value

            if not values:
                continue
            row_type = _classify_row(values)
            if row_type == "footer":
                continue

            parsed.append(SpreadsheetRow(
                sheet_name=sheet.title,
                row_number=offset,
                headers=headers,
                values=values,
                text=_row_to_text(display_filename or path.name, sheet.title, offset, values, row_type),
                row_type=row_type,
            ))

    return parsed


def _parse_xls(path: Path, display_filename: str = None) -> List[SpreadsheetRow]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("缺少 xlrd，无法解析旧版 .xls 文件，请安装依赖或另存为 .xlsx/CSV") from exc

    workbook = xlrd.open_workbook(str(path), formatting_info=True)
    parsed: List[SpreadsheetRow] = []

    for sheet in workbook.sheets():
        rows = _xls_rows_with_merged_values(sheet, workbook.datemode)
        header_idx = _find_header_row(rows)
        if header_idx is None:
            continue

        headers = [_normalize_header(value, i) for i, value in enumerate(rows[header_idx])]
        for offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            values = {}
            for i, header in enumerate(headers):
                cell_value = row[i] if i < len(row) else None
                if _format_cell(cell_value):
                    values[header] = cell_value

            if not values:
                continue
            row_type = _classify_row(values)
            if row_type == "footer":
                continue

            parsed.append(SpreadsheetRow(
                sheet_name=sheet.name,
                row_number=offset,
                headers=headers,
                values=values,
                text=_row_to_text(display_filename or path.name, sheet.name, offset, values, row_type),
                row_type=row_type,
            ))

    return parsed


def _xls_rows_with_merged_values(sheet, datemode: int) -> List[List[Any]]:
    merged_lookup = {}
    for row_start, row_end, col_start, col_end in sheet.merged_cells:
        top_left = _xls_cell_value(sheet.cell(row_start, col_start), datemode)
        for row in range(row_start, row_end):
            for col in range(col_start, col_end):
                merged_lookup[(row, col)] = top_left

    rows = []
    for row in range(sheet.nrows):
        values = []
        for col in range(sheet.ncols):
            values.append(merged_lookup.get((row, col), _xls_cell_value(sheet.cell(row, col), datemode)))
        rows.append(values)
    return rows


def _xls_cell_value(cell, datemode: int) -> Any:
    try:
        import xlrd
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, datemode)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if cell.value else "FALSE"
        if cell.ctype == xlrd.XL_CELL_ERROR:
            return ""
    except Exception:
        pass
    return cell.value


def _sheet_rows_with_merged_values(sheet) -> List[List[Any]]:
    merged_lookup = {}
    max_row = 0
    max_col = 0
    for merged_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col).value
        if _format_cell(top_left):
            max_row = max(max_row, merged_range.max_row)
            max_col = max(max_col, merged_range.max_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row, col)] = top_left

    for (row, col), cell in sheet._cells.items():
        if _format_cell(cell.value):
            max_row = max(max_row, row)
            max_col = max(max_col, col)

    if max_row <= 0 or max_col <= 0:
        return []

    rows = []
    for row in range(1, max_row + 1):
        values = []
        for col in range(1, max_col + 1):
            values.append(merged_lookup.get((row, col), sheet.cell(row, col).value))
        rows.append(values)
    return rows


def _parse_csv(path: Path, display_filename: str = None) -> List[SpreadsheetRow]:
    rows: List[List[str]] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = [row for row in reader]

    header_idx = _find_header_row(rows)
    if header_idx is None:
        return []

    headers = [_normalize_header(value, i) for i, value in enumerate(rows[header_idx])]
    parsed: List[SpreadsheetRow] = []
    for offset, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        values = {}
        for i, header in enumerate(headers):
            cell_value = row[i] if i < len(row) else ""
            if _format_cell(cell_value):
                values[header] = cell_value

        if not values:
            continue
        row_type = _classify_row(values)
        if row_type == "footer":
            continue

        parsed.append(SpreadsheetRow(
            sheet_name="CSV",
            row_number=offset,
            headers=headers,
            values=values,
            text=_row_to_text(display_filename or path.name, "CSV", offset, values, row_type),
            row_type=row_type,
        ))

    return parsed


def build_spreadsheet_validation(rows: List[SpreadsheetRow]) -> Dict:
    """Build a compact parse-quality summary for admins."""
    if not rows:
        return {
            "ok": False,
            "row_count": 0,
            "sheet_count": 0,
            "empty_ratio": 1,
            "duplicate_rows": 0,
            "row_type_counts": {},
            "warnings": ["未解析到有效数据行"],
            "sheets": [],
            "samples": [],
        }

    total_cells = 0
    filled_cells = 0
    row_value_counts = {}
    row_type_counts: Dict[str, int] = {}
    sheet_stats: Dict[str, Dict] = {}
    warnings = []

    for row in rows:
        headers = row.headers or list(row.values.keys())
        row_type_counts[row.row_type] = row_type_counts.get(row.row_type, 0) + 1
        total_cells += len(headers)
        filled_cells += sum(1 for header in headers if _format_cell(row.values.get(header)))
        value_signature = json.dumps(
            {key: _format_cell(row.values.get(key)) for key in headers},
            ensure_ascii=False,
            sort_keys=True,
        )
        row_value_counts[value_signature] = row_value_counts.get(value_signature, 0) + 1
        sheet = sheet_stats.setdefault(row.sheet_name, {
            "sheet_name": row.sheet_name,
            "row_count": 0,
            "columns": [],
            "column_sets": set(),
        })
        sheet["row_count"] += 1
        sheet["column_sets"].add(tuple(headers))
        for header in headers:
            if header not in sheet["columns"]:
                sheet["columns"].append(header)

    duplicate_rows = sum(count - 1 for count in row_value_counts.values() if count > 1)
    empty_ratio = 1 - (filled_cells / total_cells) if total_cells else 1
    inconsistent_sheets = [
        sheet["sheet_name"]
        for sheet in sheet_stats.values()
        if len(sheet["column_sets"]) > 1
    ]
    if empty_ratio > 0.35:
        warnings.append(f"空单元格比例较高: {empty_ratio:.0%}")
    if duplicate_rows:
        warnings.append(f"存在 {duplicate_rows} 行重复数据")
    if inconsistent_sheets:
        warnings.append("部分 Sheet 列结构不一致: " + "、".join(inconsistent_sheets[:5]))
    if row_type_counts.get("summary"):
        warnings.append(f"识别到 {row_type_counts['summary']} 行汇总数据")

    sheets = []
    for sheet in sheet_stats.values():
        sheets.append({
            "sheet_name": sheet["sheet_name"],
            "row_count": sheet["row_count"],
            "columns": sheet["columns"][:30],
            "column_count": len(sheet["columns"]),
            "consistent_columns": len(sheet["column_sets"]) == 1,
        })

    samples = []
    for row in rows[:3]:
        samples.append({
            "sheet_name": row.sheet_name,
            "row_number": row.row_number,
            "row_type": row.row_type,
            "values": {
                key: _format_cell(value)
                for key, value in list(row.values.items())[:8]
            },
        })

    return {
        "ok": not warnings,
        "row_count": len(rows),
        "sheet_count": len(sheet_stats),
        "empty_ratio": round(empty_ratio, 4),
        "duplicate_rows": duplicate_rows,
        "row_type_counts": row_type_counts,
        "warnings": warnings,
        "sheets": sheets,
        "samples": samples,
    }


class SpreadsheetStore:
    """SQLite store for exact spreadsheet lookup and validation."""

    def __init__(self, db_path: Union[str, Path]):
        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    def _connect(self):
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self):
        with self._connect() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS spreadsheet_files (
                    content_hash TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    source_path TEXT NOT NULL,
                    category TEXT NOT NULL,
                    access_level TEXT NOT NULL,
                    department TEXT DEFAULT '',
                    uploaded_by TEXT NOT NULL,
                    uploaded_at TEXT NOT NULL,
                    validation_json TEXT DEFAULT '{}'
                )
            """)
            file_columns = {
                row["name"]
                for row in conn.execute("PRAGMA table_info(spreadsheet_files)").fetchall()
            }
            if "validation_json" not in file_columns:
                conn.execute("ALTER TABLE spreadsheet_files ADD COLUMN validation_json TEXT DEFAULT '{}'")
            conn.execute("""
                CREATE TABLE IF NOT EXISTS spreadsheet_rows (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    content_hash TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    sheet_name TEXT NOT NULL,
                    row_number INTEGER NOT NULL,
                    row_type TEXT DEFAULT 'data',
                    headers_json TEXT NOT NULL,
                    values_json TEXT NOT NULL,
                    row_text TEXT NOT NULL,
                    category TEXT NOT NULL,
                    access_level TEXT NOT NULL,
                    department TEXT DEFAULT '',
                    source_path TEXT NOT NULL,
                    uploaded_by TEXT NOT NULL,
                    uploaded_at TEXT NOT NULL
                )
            """)
            row_columns = {
                row["name"]
                for row in conn.execute("PRAGMA table_info(spreadsheet_rows)").fetchall()
            }
            if "row_type" not in row_columns:
                conn.execute("ALTER TABLE spreadsheet_rows ADD COLUMN row_type TEXT DEFAULT 'data'")
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_spreadsheet_rows_source
                ON spreadsheet_rows(content_hash, sheet_name, row_number)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_spreadsheet_rows_acl
                ON spreadsheet_rows(category, access_level, department)
            """)
            conn.commit()

    def upsert_file_rows(self, *, content_hash: str, filename: str, source_path: str,
                         category: str, access_level: str, department: str,
                         uploaded_by: str, uploaded_at: str,
                         rows: List[SpreadsheetRow]) -> int:
        with self._connect() as conn:
            validation = build_spreadsheet_validation(rows)
            conn.execute("""
                INSERT OR REPLACE INTO spreadsheet_files
                (content_hash, filename, source_path, category, access_level,
                 department, uploaded_by, uploaded_at, validation_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                content_hash, filename, source_path, category, access_level,
                department, uploaded_by, uploaded_at,
                json.dumps(validation, ensure_ascii=False),
            ))
            conn.execute(
                "DELETE FROM spreadsheet_rows WHERE content_hash = ?",
                (content_hash,),
            )
            for row in rows:
                conn.execute("""
                    INSERT INTO spreadsheet_rows
                    (content_hash, filename, sheet_name, row_number, row_type, headers_json,
                     values_json, row_text, category, access_level, department,
                     source_path, uploaded_by, uploaded_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    content_hash,
                    filename,
                    row.sheet_name,
                    row.row_number,
                    row.row_type,
                    json.dumps(row.headers, ensure_ascii=False),
                    json.dumps(row.values, ensure_ascii=False, default=_format_cell),
                    row.text,
                    category,
                    access_level,
                    department,
                    source_path,
                    uploaded_by,
                    uploaded_at,
                ))
            conn.commit()
        return len(rows)

    def get_rows_by_source(self, content_hash: str, sheet_name: Optional[str] = None,
                           row_start: Optional[int] = None,
                           row_end: Optional[int] = None) -> List[Dict]:
        query = "SELECT * FROM spreadsheet_rows WHERE content_hash = ?"
        params: List[Any] = [content_hash]
        if sheet_name:
            query += " AND sheet_name = ?"
            params.append(sheet_name)
        if row_start is not None:
            query += " AND row_number >= ?"
            params.append(row_start)
        if row_end is not None:
            query += " AND row_number <= ?"
            params.append(row_end)
        query += " ORDER BY sheet_name, row_number"

        with self._connect() as conn:
            rows = conn.execute(query, params).fetchall()
        return [self._row_to_dict(row) for row in rows]

    def delete_file(self, content_hash: str) -> None:
        with self._connect() as conn:
            conn.execute(
                "DELETE FROM spreadsheet_rows WHERE content_hash = ?",
                (content_hash,),
            )
            conn.execute(
                "DELETE FROM spreadsheet_files WHERE content_hash = ?",
                (content_hash,),
            )
            conn.commit()

    def snapshot_file(self, content_hash: str) -> Dict:
        with self._connect() as conn:
            file_row = conn.execute(
                "SELECT * FROM spreadsheet_files WHERE content_hash = ?",
                (content_hash,),
            ).fetchone()
            row_rows = conn.execute(
                "SELECT * FROM spreadsheet_rows WHERE content_hash = ? ORDER BY id",
                (content_hash,),
            ).fetchall()
        return {
            "file": dict(file_row) if file_row else None,
            "rows": [dict(row) for row in row_rows],
        }

    def restore_file_snapshot(self, snapshot: Dict) -> None:
        if snapshot is None:
            return
        file_row = snapshot.get("file")
        rows = snapshot.get("rows", [])
        content_hash = (file_row or rows[0] if rows else {}).get("content_hash")
        if not content_hash:
            return

        with self._connect() as conn:
            conn.execute("DELETE FROM spreadsheet_rows WHERE content_hash = ?", (content_hash,))
            conn.execute("DELETE FROM spreadsheet_files WHERE content_hash = ?", (content_hash,))
            if file_row:
                conn.execute("""
                    INSERT INTO spreadsheet_files
                    (content_hash, filename, source_path, category, access_level,
                     department, uploaded_by, uploaded_at, validation_json)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    file_row["content_hash"],
                    file_row["filename"],
                    file_row["source_path"],
                    file_row["category"],
                    file_row["access_level"],
                    file_row.get("department", ""),
                    file_row["uploaded_by"],
                    file_row["uploaded_at"],
                    file_row.get("validation_json", "{}"),
                ))
            for row in rows:
                conn.execute("""
                    INSERT INTO spreadsheet_rows
                    (content_hash, filename, sheet_name, row_number, row_type, headers_json,
                     values_json, row_text, category, access_level, department,
                     source_path, uploaded_by, uploaded_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    row["content_hash"],
                    row["filename"],
                    row["sheet_name"],
                    row["row_number"],
                    row.get("row_type", "data"),
                    row["headers_json"],
                    row["values_json"],
                    row["row_text"],
                    row["category"],
                    row["access_level"],
                    row.get("department", ""),
                    row["source_path"],
                    row["uploaded_by"],
                    row["uploaded_at"],
                ))
            conn.commit()

    def update_file_metadata(self, content_hash: str, *, category: str,
                             access_level: str, department: str) -> int:
        with self._connect() as conn:
            file_cur = conn.execute("""
                UPDATE spreadsheet_files
                SET category = ?, access_level = ?, department = ?
                WHERE content_hash = ?
            """, (category, access_level, department, content_hash))
            conn.execute("""
                UPDATE spreadsheet_rows
                SET category = ?, access_level = ?, department = ?
                WHERE content_hash = ?
            """, (category, access_level, department, content_hash))
            conn.commit()
        return file_cur.rowcount

    def find_cells(self, *, value: str, content_hash: Optional[str] = None,
                   filename: Optional[str] = None, sheet_name: Optional[str] = None,
                   column_name: Optional[str] = None,
                   access_filter: Optional[Dict] = None) -> List[Dict]:
        """Find exact cell values for answer verification."""
        value = _format_cell(value)
        rows = self.get_rows_by_source(content_hash) if content_hash else self._all_rows()
        matches = []
        for row in rows:
            if not self._match_access(row, access_filter):
                continue
            if filename and row.get("filename") != filename:
                continue
            if sheet_name and row.get("sheet_name") != sheet_name:
                continue
            for header, cell_value in row.get("values", {}).items():
                if column_name and header != column_name:
                    continue
                if _format_cell(cell_value) == value:
                    matches.append({
                        "filename": row.get("filename"),
                        "sheet_name": row.get("sheet_name"),
                        "row_number": row.get("row_number"),
                        "column_name": header,
                        "value": value,
                        "source_path": row.get("source_path"),
                        "content_hash": row.get("content_hash"),
                    })
        return matches

    def list_files(self, access_filter: Optional[Dict] = None, limit: int = 100) -> List[Dict]:
        with self._connect() as conn:
            rows = conn.execute("""
                SELECT f.*,
                       COUNT(r.id) AS row_count,
                       COUNT(DISTINCT r.sheet_name) AS sheet_count
                FROM spreadsheet_files f
                LEFT JOIN spreadsheet_rows r ON r.content_hash = f.content_hash
                GROUP BY f.content_hash
                ORDER BY f.uploaded_at DESC
                LIMIT ?
            """, (int(limit),)).fetchall()
        files = [dict(row) for row in rows]
        for row in files:
            try:
                row["validation"] = json.loads(row.pop("validation_json", "{}") or "{}")
            except json.JSONDecodeError:
                row["validation"] = {}
        return [row for row in files if self._match_access(row, access_filter)]

    def query_rows(self, *, keyword: str = "", content_hash: Optional[str] = None,
                   filename: Optional[str] = None, sheet_name: Optional[str] = None,
                   column_name: Optional[str] = None, cell_value: Optional[str] = None,
                   access_filter: Optional[Dict] = None, limit: int = 50) -> List[Dict]:
        rows = self.get_rows_by_source(content_hash) if content_hash else self._all_rows()
        keyword_norm = _format_cell(keyword)
        cell_norm = _format_cell(cell_value)
        matched = []
        for row in rows:
            if len(matched) >= limit:
                break
            if not self._match_access(row, access_filter):
                continue
            if filename and row.get("filename") != filename:
                continue
            if sheet_name and row.get("sheet_name") != sheet_name:
                continue
            if keyword_norm and keyword_norm not in row.get("row_text", ""):
                continue
            if column_name or cell_norm:
                values = row.get("values", {})
                if column_name and column_name not in values:
                    continue
                search_values = [values.get(column_name)] if column_name else values.values()
                if cell_norm and not any(_format_cell(value) == cell_norm for value in search_values):
                    continue
            matched.append(row)
        return matched

    def _all_rows(self) -> List[Dict]:
        with self._connect() as conn:
            rows = conn.execute(
                "SELECT * FROM spreadsheet_rows ORDER BY filename, sheet_name, row_number"
            ).fetchall()
        return [self._row_to_dict(row) for row in rows]

    @staticmethod
    def _match_access(row: Dict, access_filter: Optional[Dict]) -> bool:
        if not access_filter:
            return True
        allowed_levels = access_filter.get("access_level", [])
        if allowed_levels and row.get("access_level") not in allowed_levels:
            return False
        if row.get("access_level") == "public":
            return True
        allowed_depts = access_filter.get("department", [])
        if allowed_depts and row.get("department", "") not in allowed_depts:
            return False
        return True

    @staticmethod
    def _row_to_dict(row) -> Dict:
        data = dict(row)
        data["headers"] = json.loads(data.pop("headers_json"))
        data["values"] = json.loads(data.pop("values_json"))
        return data
