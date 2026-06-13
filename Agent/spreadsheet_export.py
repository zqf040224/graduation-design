"""
Excel export helpers for generated text and structured spreadsheet rows.
"""

import re
from collections import OrderedDict, defaultdict
from datetime import datetime
from typing import Dict, Iterable, List, Tuple


INVALID_SHEET_CHARS = re.compile(r"[\[\]\:\*\?\/\\]")
MARKDOWN_SEPARATOR = re.compile(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$")


def safe_filename_stem(value: str, default: str = "导出表格") -> str:
    stem = re.sub(r"[\\/:*?\"<>|\r\n\t]+", "", (value or "").strip())
    stem = re.sub(r"\s+", "_", stem)
    return (stem[:40] or default)


def _safe_sheet_title(value: str, fallback: str) -> str:
    title = INVALID_SHEET_CHARS.sub("_", (value or "").strip())[:31]
    return title or fallback


def _unique_sheet_title(existing: set, title: str) -> str:
    candidate = title[:31] or "Sheet"
    if candidate not in existing:
        existing.add(candidate)
        return candidate

    base = candidate[:27]
    index = 2
    while True:
        candidate = f"{base}_{index}"[:31]
        if candidate not in existing:
            existing.add(candidate)
            return candidate
        index += 1


def _format_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _excel_safe_text(value) -> str:
    text = _format_cell(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _style_sheet(ws, header_row: int = 1) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="DDEBFF")
    header_font = Font(name="Microsoft YaHei", bold=True, color="17304D")
    body_font = Font(name="Microsoft YaHei", color="17304D")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 8
        for cell in column_cells[:200]:
            max_len = max(max_len, len(_format_cell(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions


def _new_workbook():
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法导出 Excel 文件，请先安装依赖") from exc

    wb = Workbook()
    wb.remove(wb.active)
    return wb


def workbook_from_structured_rows(rows: List[Dict]):
    wb = _new_workbook()
    existing_titles = set()

    grouped: Dict[str, List[Dict]] = defaultdict(list)
    for row in rows:
        grouped[row.get("sheet_name") or "Sheet"].append(row)

    info = wb.create_sheet(_unique_sheet_title(existing_titles, "导出说明"))
    first = rows[0] if rows else {}
    info.append(["项目", "内容"])
    info.append(["文件名", first.get("filename", "")])
    info.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["结构化行数", len(rows)])
    info.append(["分类", first.get("category", "")])
    info.append(["访问级别", first.get("access_level", "")])
    info.append(["部门", first.get("department", "")])
    _style_sheet(info)

    for sheet_name, sheet_rows in grouped.items():
        title = _unique_sheet_title(existing_titles, _safe_sheet_title(sheet_name, "数据"))
        ws = wb.create_sheet(title)

        headers = OrderedDict()
        for row in sheet_rows:
            for header in row.get("headers", []) or row.get("values", {}).keys():
                headers.setdefault(header, None)

        export_headers = ["原始行号", "行类型", *headers.keys()]
        ws.append(export_headers)
        for row in sheet_rows:
            values = row.get("values", {}) or {}
            ws.append([
                row.get("row_number", ""),
                row.get("row_type", "data"),
                *[_excel_safe_text(values.get(header)) for header in headers.keys()],
            ])
        _style_sheet(ws)

    return wb


def _split_markdown_row(line: str) -> List[str]:
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [cell.strip().replace("\\|", "|") for cell in stripped.split("|")]


def extract_markdown_tables(text: str) -> List[Tuple[List[str], List[List[str]]]]:
    lines = (text or "").splitlines()
    tables: List[Tuple[List[str], List[List[str]]]] = []
    index = 0
    while index < len(lines) - 1:
        current = lines[index]
        next_line = lines[index + 1]
        if "|" not in current or not MARKDOWN_SEPARATOR.match(next_line):
            index += 1
            continue

        headers = _split_markdown_row(current)
        body = []
        index += 2
        while index < len(lines) and "|" in lines[index] and lines[index].strip():
            row = _split_markdown_row(lines[index])
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            body.append(row[:len(headers)])
            index += 1
        tables.append((headers, body))
    return tables


def workbook_from_text(content: str):
    wb = _new_workbook()
    existing_titles = set()
    tables = extract_markdown_tables(content)

    if tables:
        for idx, (headers, rows) in enumerate(tables, start=1):
            title = _unique_sheet_title(existing_titles, f"表格{idx}")
            ws = wb.create_sheet(title)
            ws.append([_excel_safe_text(header) for header in headers])
            for row in rows:
                ws.append([_excel_safe_text(cell) for cell in row])
            _style_sheet(ws)
        return wb

    ws = wb.create_sheet(_unique_sheet_title(existing_titles, "文本内容"))
    ws.append(["序号", "内容"])
    lines = [line.strip() for line in (content or "").splitlines() if line.strip()]
    for idx, line in enumerate(lines, start=1):
        ws.append([idx, _excel_safe_text(line)])
    _style_sheet(ws)
    return wb


def workbook_to_bytes(workbook) -> bytes:
    from io import BytesIO

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
